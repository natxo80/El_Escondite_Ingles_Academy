from PySide2.QtGui import QIcon
from PySide2.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton,
                               QLineEdit, QMessageBox, QHeaderView, QAbstractScrollArea, QComboBox, QFileDialog, QLabel)
from db.database import insert_student, fetch_students, delete_student, update_student
from db.database import fetch_levels # Necesario para el como de niveles de alumnos
import csv
import openpyxl

class ManageStudentsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestionar Alumnos")  # Establece el título de la ventana
        self.setWindowIcon(QIcon("resources/el_escondite_ingles.bmp"))
        self.setGeometry(100, 100, 900, 900)      # Tamaño y posición de la ventana
        self.init_ui()                            # Carga los componentes gráficos

    def init_ui(self):
        """Configura los elementos visuales de la ventana"""
        self.layout = QVBoxLayout()  # Layout vertical principal
        self.layout.setContentsMargins(30, 30, 30, 30)  # Márgenes exteriores
        self.layout.setSpacing(15)  # Espaciado entre secciones

        # ---------- CABECERA VISUAL (Logo + Título) ----------
        header_layout = QHBoxLayout()

        self.logo = QLabel()
        self.logo.setPixmap("resources/el_escondite_ingles.jpg")  # Asegúrate de que la ruta es correcta
        self.logo.setFixedSize(60, 60)
        self.logo.setScaledContents(True)
        header_layout.addWidget(self.logo)

        title = QLabel("Gestión de Alumnos - El Escondite Inglés")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4B0082;")  # Morado elegante
        header_layout.addWidget(title)
        header_layout.addStretch()
        self.layout.addLayout(header_layout)
        # -----------------------------------------------------

        # Campo de búsqueda y botón limpiar
        search_layout = QHBoxLayout()
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Buscar alumno por nombre...")
        self.search_input.textChanged.connect(self.filter_students)
        search_layout.addWidget(self.search_input)

        self.clear_search_button = QPushButton("Limpiar filtro")
        self.clear_search_button.clicked.connect(self.clear_search)
        search_layout.addWidget(self.clear_search_button)

        self.layout.addLayout(search_layout)

        # Tabla
        self.table = QTableWidget()
        self.table.setColumnCount(3)
        self.table.setHorizontalHeaderLabels(["Nombre", "Edad", "Nivel"])
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)
        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)
        self.layout.addWidget(self.table)

        # Formulario de entrada
        form_layout = QHBoxLayout()
        self.name_input = QLineEdit()
        self.name_input.setPlaceholderText("Nombre del alumno")
        form_layout.addWidget(self.name_input)

        self.age_input = QLineEdit()
        self.age_input.setPlaceholderText("Edad")
        form_layout.addWidget(self.age_input)

        self.level_combo = QComboBox()
        self.load_levels_into_combo()
        form_layout.addWidget(self.level_combo)
        self.layout.addLayout(form_layout)

        # Botones de acción (Agregar, Editar, Borrar)
        button_layout = QHBoxLayout()
        self.add_button = QPushButton("Añadir Alumno")
        self.add_button.clicked.connect(self.add_student)
        button_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Editar Alumno")
        self.edit_button.clicked.connect(self.edit_student)
        button_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Borrar Alumno")
        self.delete_button.clicked.connect(self.delete_student)
        button_layout.addWidget(self.delete_button)

        self.layout.addLayout(button_layout)

        # Botones de exportación
        export_layout = QHBoxLayout()
        self.export_csv_button = QPushButton("Exportar CSV")
        self.export_csv_button.clicked.connect(self.export_to_csv)
        export_layout.addWidget(self.export_csv_button)

        self.export_excel_button = QPushButton("Exportar Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_excel_button)

        self.layout.addLayout(export_layout)
        self.setLayout(self.layout)

        self.load_students()  # Carga inicial de alumnos en la tabla

    def clear_search(self):
        """Limpia el campo de búsqueda y recarga todos los alumnos"""
        self.search_input.clear()
        self.load_students()

    def load_levels_into_combo(self):
        """Carga todos los niveles desde la base de datos al ComboBox"""
        self.level_combo.clear()
        levels = fetch_levels()
        for level in levels:
            self.level_combo.addItem(level[1])

    def load_students(self, filter_text=""):
        """
        Carga los alumnos desde la base de datos y los muestra en la tabla.
        Si se pasa un filtro, solo muestra los alumnos cuyo nombre lo contenga.
        """
        self.table.setRowCount(0)  # Limpiamos la tabla antes de llenarla
        students = fetch_students()         # Obtenemos todos los alumnos

        # Si hay texto en el filtro, aplicamos búsqueda por nombre
        if filter_text:
            filter_text = filter_text.lower()
            students = [s for s in students if filter_text in s["name"].lower()]

        # Llenamos la tabla con los datos (filtrados o completos)
        for row, student in enumerate(students):
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(student["name"]))
            self.table.setItem(row, 1, QTableWidgetItem(str(student["age"])))
            self.table.setItem(row, 2, QTableWidgetItem(student["level"]))

    def add_student(self):
        """Añade un nuevo alumno usando los campos de texto"""
        name = self.name_input.text().strip()
        age = self.age_input.text().strip()
        level = self.level_combo.currentText()

        # Validación: campos obligatorios
        if not name or not age or not level:
            QMessageBox.warning(self, "Error", "Todos los campos son obligatorios.")
            return

        # Validación: edad debe ser un número válido entre 12 y 120
        if not age.isdigit() or not (12 <= int(age) <= 120):
            QMessageBox.warning(self, "Error", "La edad debe estar entre 12 y 120.")
            return

        # Validación: no permitir duplicados por nombre
        existing_names = [self.table.item(row, 0).text().lower() for row in range(self.table.rowCount())]
        if name.lower() in existing_names:
            QMessageBox.warning(self, "Error", f"El alumno '{name}' ya existe.")
            return

        # Si pasa todas las validaciones, insertamos y limpiamos campos
        insert_student(name, int(age), level)
        self.load_students()
        self.name_input.clear()
        self.age_input.clear()
        self.level_combo.setCurrentIndex(0)

    def edit_student(self):
        """Edita un alumno seleccionado usando los nuevos datos de los campos"""
        selected = self.table.currentRow()
        if selected != -1:
            name = self.name_input.text().strip()
            age = self.age_input.text().strip()
            level = self.level_combo.currentText()
            old_name = self.table.item(selected, 0).text()  # Nombre original

            if name and age.isdigit() and level:
                update_student(old_name, name, int(age), level)
                self.load_students()
                self.name_input.clear()
                self.age_input.clear()
                self.level_combo.setCurrentIndex(0)
            else:
                QMessageBox.warning(self, "Error", "Todos los campos deben ser válidos.")
        else:
            QMessageBox.warning(self, "Error", "Selecciona un alumno para editar.")

    def delete_student(self):
        """Elimina el alumno seleccionado de la tabla y la base de datos con validaciones"""
        selected = self.table.currentRow()

        if selected == -1:
            QMessageBox.warning(self, "Error", "Selecciona un alumno para borrar.")
            return

        name = self.table.item(selected, 0).text()

        # Confirmación antes de borrar
        confirm = QMessageBox.question(
            self,
            "Confirmar borrado",
            f"¿Estás seguro de que deseas eliminar al alumno '{name}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.Yes:
            try:
                delete_student(name)
                self.load_students()
                QMessageBox.information(self, "Alumno eliminado", f"'{name}' ha sido eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error al borrar", f"No se pudo borrar al alumno:\n{str(e)}")

    def filter_students(self):
        """
        Se llama automáticamente cada vez que el usuario escribe en el campo de búsqueda.
        Actualiza la tabla para mostrar solo los alumnos que coincidan con el texto.
        """
        search_text = self.search_input.text()  # Obtenemos el texto de búsqueda
        self.load_students(search_text)  # Recargamos la tabla con el filtro aplicado

    def filter_students(self):
        """Filtra la tabla de alumnos por nombre"""
        search_text = self.search_input.text()
        self.load_students(search_text)

    def export_to_csv(self):
          """Exporta los alumnos a un archivo CSV"""
          # Abrir diálogo para elegir dónde guardar el archivo
          path, _ = QFileDialog.getSaveFileName(self, "Guardar como CSV", "alumnos.csv", "CSV Files (*.csv)")
          if not path:
              return

          # Por si se cambia el nombre y no se le pone la extension .csv
          if not path.endswith(".csv"):
              path += ".csv"

          with open(path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow([self.table.horizontalHeaderItem(i).text() for i in range(self.table.columnCount())])
            for row in range(self.table.rowCount()):
                 writer.writerow([self.table.item(row, col).text() for col in range(self.table.columnCount())])

          QMessageBox.information(self, "Exportación exitosa", f"Datos exportados a:{path}")

    def export_to_excel(self):
         """Exporta los alumnos a un archivo Excel (.xlsx) usando openpyxl"""
         # Abrir diálogo para elegir dónde guardar el archivo
         path, _ = QFileDialog.getSaveFileName(self, "Guardar como Excel", "alumnos.xlsx", "Archivos Excel (*.xlsx)")
         if not path:
             return  # Cancelado por el usuario

         # Por si se cambia el nombre y no se le pone la extension .xlsx
         if not path.endswith(".xlsx"):
             path += ".xlsx"

         workbook = openpyxl.Workbook()
         sheet = workbook.active
         sheet.title = "Alumnos"

         # Escribir cabeceras
         headers = ["Nombre", "Edad", "Nivel"]
         for col, header in enumerate(headers, start=1):  # Empieza en columna 1
             sheet.cell(row=1, column=col, value=header)

        # Escribir datos de la tabla
         for row in range(self.table.rowCount()):
             for col in range(self.table.columnCount()):
                item = self.table.item(row, col)
                if item:
                    sheet.cell(row=row + 2, column=col + 1, value=item.text())  # +2 porque empieza en fila 2

         workbook.save(path)
         QMessageBox.information(self, "Éxito", f"Archivo Excel exportado correctamente:\n{path}")
