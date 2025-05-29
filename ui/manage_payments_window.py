from PySide2.QtGui import QIcon
from PySide2.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton,
                               QLineEdit,
                               QMessageBox, QHeaderView, QAbstractScrollArea, QAbstractItemView, QComboBox, QFileDialog,
                               QLabel)
from db.database import fetch_payments, insert_payment, update_payment, delete_payment
from db.database import fetch_students  # Necesario para obtener ID de alumnos
from db.database import is_student_under_bonus # Necesario para saber si un estudiante tiene un bono activo
import csv
import openpyxl


class ManagePaymentsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Gestionar Pagos")  # Título de la ventana
        self.setWindowIcon(QIcon("resources/el_escondite_ingles.bmp"))
        self.setGeometry(100, 100, 1400, 800)     # Tamaño y posición
        self.init_ui()

    def init_ui(self):
        """Inicializa la interfaz gráfica de gestión de pagos con diseño visual mejorado"""
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(30, 30, 30, 30)  # Márgenes exteriores
        self.layout.setSpacing(15)  # Espaciado vertical entre secciones

        # ---------- CABECERA VISUAL ----------
        header_layout = QHBoxLayout()

        self.logo = QLabel()
        self.logo.setPixmap("resources/el_escondite_ingles.jpg")  # Ruta al logo de la academia
        self.logo.setFixedSize(60, 60)
        self.logo.setScaledContents(True)
        header_layout.addWidget(self.logo)

        title = QLabel("Gestión de Pagos - El Escondite Inglés")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4B0082;")
        header_layout.addWidget(title)
        header_layout.addStretch()

        self.layout.addLayout(header_layout)
        # -------------------------------------

        # ---------- FILTROS DE BÚSQUEDA ----------
        search_layout = QHBoxLayout()

        self.search_student_input = QLineEdit()
        self.search_student_input.setPlaceholderText("Buscar por alumno")
        self.search_student_input.textChanged.connect(self.load_payments)
        search_layout.addWidget(self.search_student_input)

        self.search_notes_input = QLineEdit()
        self.search_notes_input.setPlaceholderText("Buscar por notas")
        self.search_notes_input.textChanged.connect(self.load_payments)
        search_layout.addWidget(self.search_notes_input)

        clear_filters_button = QPushButton("Limpiar Filtros")
        clear_filters_button.clicked.connect(self.clear_filters)
        search_layout.addWidget(clear_filters_button)

        self.layout.addLayout(search_layout)
        # ------------------------------------------

        # ---------- TABLA DE PAGOS ----------
        self.table = QTableWidget()
        self.table.setColumnCount(6)
        self.table.setHorizontalHeaderLabels(["ID", "Alumno", "Cantidad", "Fecha", "Método", "Notas"])

        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.table.setHorizontalScrollMode(QAbstractItemView.ScrollPerPixel)
        self.table.setVerticalScrollMode(QAbstractItemView.ScrollPerPixel)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)

        self.layout.addWidget(self.table)
        # -----------------------------------

        # ---------- FORMULARIO DE ENTRADA ----------
        form_layout = QHBoxLayout()

        self.student_combo = QComboBox()
        self.load_students()
        form_layout.addWidget(self.student_combo)

        self.amount_input = QLineEdit()
        self.amount_input.setPlaceholderText("Cantidad (€)")
        form_layout.addWidget(self.amount_input)

        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("Fecha (YYYY-MM-DD)")
        form_layout.addWidget(self.date_input)

        self.method_input = QLineEdit()
        self.method_input.setPlaceholderText("Método (efectivo, tarjeta...)")
        form_layout.addWidget(self.method_input)

        self.notes_input = QLineEdit()
        self.notes_input.setPlaceholderText("Notas (opcional)")
        form_layout.addWidget(self.notes_input)

        self.layout.addLayout(form_layout)
        # ------------------------------------------

        # ---------- BOTONES DE ACCIÓN ----------
        button_layout = QHBoxLayout()

        self.add_button = QPushButton("Añadir Pago")
        self.add_button.clicked.connect(self.add_payment)
        button_layout.addWidget(self.add_button)

        self.edit_button = QPushButton("Editar Pago")
        self.edit_button.clicked.connect(self.edit_payment)
        button_layout.addWidget(self.edit_button)

        self.delete_button = QPushButton("Eliminar Pago")
        self.delete_button.clicked.connect(self.delete_payment)
        button_layout.addWidget(self.delete_button)

        self.layout.addLayout(button_layout)
        # ---------------------------------------

        # ---------- EXPORTACIÓN ----------
        export_layout = QHBoxLayout()

        self.export_csv_button = QPushButton("Exportar CSV")
        self.export_csv_button.clicked.connect(self.export_to_csv)
        export_layout.addWidget(self.export_csv_button)

        self.export_excel_button = QPushButton("Exportar Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_excel_button)

        self.layout.addLayout(export_layout)
        # ---------------------------------

        self.setLayout(self.layout)
        self.load_payments()

    def clear_filters(self):
        """Limpia los campos de búsqueda y recarga la tabla sin filtros"""
        self.search_student_input.clear()
        self.search_notes_input.clear()
        self.load_payments()

    def load_students(self):
        """Carga los alumnos en el combobox"""
        self.student_combo.clear()
        self.student_combo.addItem("Selecciona alumno", None)
        for student in fetch_students():
            label = f"{student['name']} (ID: {student['id']})"
            self.student_combo.addItem(label, student["id"])

    def load_payments(self):
        """Carga los pagos desde la base de datos y los muestra en la tabla, aplicando filtros si es necesario"""
        student_filter = self.search_student_input.text().lower()
        notes_filter = self.search_notes_input.text().lower()

        self.table.setRowCount(0)
        payments = fetch_payments()  # [(id, alumno, cantidad, fecha, metodo, notas), ...]

        for payment in payments:
            student_name = payment[1].lower()
            notes = (payment[5] or "").lower()  # Por si es None

            # Aplicar filtros
            if student_filter and student_filter not in student_name:
                continue
            if notes_filter and notes_filter not in notes:
                continue

            row = self.table.rowCount()
            self.table.insertRow(row)
            for col, value in enumerate(payment):
                self.table.setItem(row, col, QTableWidgetItem(str(value)))

    def add_payment(self):
        """Añade un nuevo pago"""
        student_id = self.student_combo.currentData() # Obtenemos el ID del alumno seleccionado en el combo desplegable
        amount = self.amount_input.text().strip()  # Cantidad pagada
        date = self.date_input.text().strip()  # Fecha del pago (formato YYYY-MM-DD)
        method = self.method_input.text().strip()  # Métod de pago (opcional)
        notes = self.notes_input.text().strip()  # Notas (opcional)

        # Validación: Asegurar que se haya seleccionado un alumno válido
        if not student_id:
            QMessageBox.warning(self, "Error", "Selecciona un alumno válido.")
            return

        # Validación: Comprobar que los campos obligatorios (cantidad y fecha) no estén vacíos
        if not amount or not date:
            QMessageBox.warning(self, "Error", "Cantidad y fecha son obligatorios.")
            return

        # Validación: Verificar que la cantidad sea un número válido y positivo
        if not amount.replace('.', '', 1).isdigit() or float(amount) <= 0:
            QMessageBox.warning(self, "Error", "La cantidad debe ser un número positivo.")
            return

        # Validación: Revisar si el alumno tiene un bono gratuito activo
        if is_student_under_bonus(student_id):
            QMessageBox.warning(self, "Error", "Este alumno está actualmente en periodo de bono gratuito.")
            return

        insert_payment(student_id, float(amount), date, method, notes)
        self.load_payments()
        self.clear_fields()

    def edit_payment(self):
        """Edita el pago seleccionado"""
        row = self.table.currentRow()  # Obtenemos la fila seleccionada en la tabla
        if row == -1:
            QMessageBox.warning(self, "Error", "Selecciona un pago para editar.")  # Si no hay selección, mostramos error
            return

        payment_id = int(self.table.item(row, 0).text())  # Obtenemos el ID del pago seleccionado
        amount = self.amount_input.text().strip()  # Obtenemos la nueva cantidad
        date = self.date_input.text().strip()  # Obtenemos la nueva fecha
        method = self.method_input.text().strip()  # Nuevo metodo (opcional)
        notes = self.notes_input.text().strip()  # Nuevas notas (opcional)

        # Validación: la cantidad y la fecha son obligatorias
        if not amount or not date:
            QMessageBox.warning(self, "Error", "Cantidad y fecha son obligatorios.")  # Mostramos error
            return

        # Validación: la cantidad debe ser un número válido mayor que 0
        if not amount.replace('.', '', 1).isdigit() or float(amount) <= 0:
            QMessageBox.warning(self, "Error", "La cantidad debe ser un número positivo.")  # Mostramos error
            return

        update_payment(payment_id, float(amount), date, method, notes)  # Actualizamos el pago en la base de datos
        self.load_payments()  # Recargamos la tabla
        self.clear_fields()  # Limpiamos los campos de entrada

    def delete_payment(self):
        """Elimina el pago seleccionado con confirmación y manejo de errores"""
        row = self.table.currentRow()

        if row == -1:
            QMessageBox.warning(self, "Error", "Selecciona un pago para eliminar.")
            return

        payment_id = int(self.table.item(row, 0).text())

        # Confirmación del usuario
        confirm = QMessageBox.question(
            self,
            "Confirmar eliminación",
            f"¿Estás seguro de que deseas eliminar el pago con ID {payment_id}?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.Yes:
            try:
                delete_payment(payment_id)
                self.load_payments()
                self.clear_fields()
                QMessageBox.information(self, "Pago eliminado",
                                        f"El pago con ID {payment_id} ha sido eliminado correctamente.")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo eliminar el pago:\n{str(e)}")

    def clear_fields(self):
        """Limpia los campos de entrada"""
        self.student_combo.setCurrentIndex(0)
        self.amount_input.clear()
        self.date_input.clear()
        self.method_input.clear()
        self.notes_input.clear()

    def export_to_csv(self):
        """Exporta los pagos a un archivo CSV"""
        # Abrir diálogo para elegir dónde guardar el archivo
        path, _ = QFileDialog.getSaveFileName(self, "Guardar como CSV", "pagos.csv", "CSV Files (*.csv)")
        if not path:
            return

        # Por si se cambia el nombre y no se le pone la extension .csv
        if not path.endswith(".csv"):
            path += ".csv"

        with open(path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["ID", "Alumno", "Cantidad", "Fecha", "Método", "Notas"])

            for row in range(self.table.rowCount()):
                row_data = [self.table.item(row, col).text() if self.table.item(row, col) else "" for col in
                            range(self.table.columnCount())]
                writer.writerow(row_data)

        QMessageBox.information(self, "Éxito", f"Archivo CSV exportado correctamente:\n{path}")

    def export_to_excel(self):
         """Exporta los pagos a un archivo Excel (.xlsx) usando openpyxl"""
         # Abrir diálogo para elegir dónde guardar el archivo
         path, _ = QFileDialog.getSaveFileName(self, "Guardar como Excel", "pagos.xlsx", "Archivos Excel (*.xlsx)")
         if not path:
             return  # Cancelado por el usuario

         # Por si se cambia el nombre y no se le pone la extension .xlsx
         if not path.endswith(".xlsx"):
             path += ".xlsx"

         workbook = openpyxl.Workbook()
         sheet = workbook.active
         sheet.title = "Pagos"

         # Escribir cabeceras
         headers = ["ID", "Alumno", "Cantidad", "Fecha", "Método", "Notas"]
         for col, header in enumerate(headers, start=1):
             sheet.cell(row=1, column=col, value=header)

        # Escribir datos de la tabla
         for row in range(self.table.rowCount()):
             for col in range(self.table.columnCount()):
                 item = self.table.item(row, col)
                 sheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "") # +2 porque empieza en fila 2

         workbook.save(path)
         QMessageBox.information(self, "Éxito", f"Archivo Excel exportado correctamente:\n{path}")