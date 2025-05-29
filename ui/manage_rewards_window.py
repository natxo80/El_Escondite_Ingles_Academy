from PySide2.QtGui import QIcon
from PySide2.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QTableWidget, QTableWidgetItem, QPushButton,
                               QMessageBox,
                               QLabel, QHeaderView, QComboBox, QLineEdit, QFileDialog, QAbstractScrollArea)
from PySide2.QtCore import Qt
from datetime import datetime
from db.database import fetch_rewards, fetch_students, insert_reward, reward_already_granted
from db.database import get_students_with_expiring_bonus # Para saber próximas finalizaciones de bonos
import csv
import openpyxl

class ManageRewardsWindow(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Recompensas por Recomendación")  # Título de la ventana
        self.setWindowIcon(QIcon("resources/el_escondite_ingles.bmp"))
        self.setGeometry(100, 100, 1200, 800)  # Tamaño de la ventana
        self.init_ui()  # Cargar interfaz

    def init_ui(self):
        """Inicializa los elementos visuales de la ventana de recompensas con diseño estético mejorado"""
        self.layout = QVBoxLayout()
        self.layout.setContentsMargins(30, 30, 30, 30)
        self.layout.setSpacing(15)

        # ---------- CABECERA VISUAL ----------
        header_layout = QHBoxLayout()

        self.logo = QLabel()
        self.logo.setPixmap("resources/el_escondite_ingles.jpg")  # Asegúrate de tener esta imagen en la ruta correcta
        self.logo.setFixedSize(60, 60)
        self.logo.setScaledContents(True)
        header_layout.addWidget(self.logo)

        title = QLabel("Gestión de Recompensas - El Escondite Inglés")
        title.setStyleSheet("font-size: 20px; font-weight: bold; color: #4B0082;")
        header_layout.addWidget(title)
        header_layout.addStretch()

        self.layout.addLayout(header_layout)
        # -------------------------------------

        # ---------- FILTROS DE BÚSQUEDA ----------
        filter_layout = QHBoxLayout()

        self.search_recommender = QLineEdit()
        self.search_recommender.setPlaceholderText("Buscar por recomendador")
        self.search_recommender.textChanged.connect(self.load_rewards)
        filter_layout.addWidget(self.search_recommender)

        self.search_new_student = QLineEdit()
        self.search_new_student.setPlaceholderText("Buscar por nuevo alumno")
        self.search_new_student.textChanged.connect(self.load_rewards)
        filter_layout.addWidget(self.search_new_student)

        self.search_reward_name = QLineEdit()
        self.search_reward_name.setPlaceholderText("Buscar por nombre de recompensa")
        self.search_reward_name.textChanged.connect(self.load_rewards)
        filter_layout.addWidget(self.search_reward_name)

        clear_filter_button = QPushButton("Limpiar Filtros")
        clear_filter_button.clicked.connect(self.clear_filters)
        filter_layout.addWidget(clear_filter_button)

        self.layout.addLayout(filter_layout)
        # ------------------------------------------

        # ---------- TABLA DE RECOMPENSAS ----------
        self.table = QTableWidget()
        self.table.setColumnCount(5)
        self.table.setHorizontalHeaderLabels([
            "Recomendador", "Alumno nuevo", "Nombre de la recompensa",
            "Meses de recompensa", "Fecha de inicio bono"
        ])
        self.table.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
        self.table.setHorizontalScrollMode(QTableWidget.ScrollPerPixel)
        self.table.setVerticalScrollMode(QTableWidget.ScrollPerPixel)

        header = self.table.horizontalHeader()
        header.setSectionResizeMode(QHeaderView.Stretch)
        header.setStretchLastSection(True)

        self.layout.addWidget(self.table)
        # ------------------------------------------

        # ---------- FORMULARIO DE INSERCIÓN ----------
        form_layout = QHBoxLayout()

        self.recommender_combo = QComboBox()
        self.recommender_combo.setPlaceholderText("Selecciona quien recomienda")
        form_layout.addWidget(self.recommender_combo)

        self.new_student_combo = QComboBox()
        self.new_student_combo.setPlaceholderText("Selecciona el nuevo alumno")
        form_layout.addWidget(self.new_student_combo)

        self.months_input = QLineEdit()
        self.months_input.setPlaceholderText("Meses de recompensa (1, 3, 6)")
        form_layout.addWidget(self.months_input)

        self.date_input = QLineEdit()
        self.date_input.setPlaceholderText("Fecha de inicio bono (YYYY-MM-DD)")
        form_layout.addWidget(self.date_input)
        
        self.layout.addLayout(form_layout)
        # ---------------------------------------------

        # ---------- BOTONES: Añadir / Eliminar recompensa ----------
        action_buttons_layout = QHBoxLayout()

        self.add_button = QPushButton("Añadir Recompensa")
        self.add_button.clicked.connect(self.add_reward)
        action_buttons_layout.addWidget(self.add_button)

        self.delete_button = QPushButton("Borrar Recompensa")
        self.delete_button.clicked.connect(self.delete_reward)
        action_buttons_layout.addWidget(self.delete_button)

        self.layout.addLayout(action_buttons_layout)
        # -----------------------------------------------------------------------

        # ---------- BOTONES DE UTILIDAD ----------
        refresh_button = QPushButton("Actualizar Recompensas")
        refresh_button.clicked.connect(self.load_rewards)
        self.layout.addWidget(refresh_button)

        check_expiring_button = QPushButton("Ver Bonos que Finalizan Pronto")
        check_expiring_button.clicked.connect(self.show_expiring_alerts)
        self.layout.addWidget(check_expiring_button)
        # ------------------------------------------

        # ---------- BOTONES DE EXPORTACIÓN ----------
        export_layout = QHBoxLayout()

        self.export_csv_button = QPushButton("Exportar CSV")
        self.export_csv_button.clicked.connect(self.export_to_csv)
        export_layout.addWidget(self.export_csv_button)

        self.export_excel_button = QPushButton("Exportar Excel")
        self.export_excel_button.clicked.connect(self.export_to_excel)
        export_layout.addWidget(self.export_excel_button)

        self.layout.addLayout(export_layout)
        # --------------------------------------------

        self.setLayout(self.layout)

        self.load_students()
        self.load_rewards()

    def load_students(self):
        """Carga todos los alumnos en los combos para selección"""
        students = fetch_students()
        self.recommender_combo.clear()
        self.new_student_combo.clear()
        for student in students:
            label = f"{student['name']} (ID: {student['id']})"
            self.recommender_combo.addItem(label, student["id"])
            self.new_student_combo.addItem(label, student["id"])

    def load_rewards(self):
        """Carga las recompensas desde la base de datos y las muestra en la tabla"""
        self.table.setRowCount(0)  # Limpiar tabla
        rewards = fetch_rewards()  # Obtener recompensas

        # Aplicar filtros
        filter_recommender = self.search_recommender.text().lower()
        filter_new_student = self.search_new_student.text().lower()
        filter_reward_name = self.search_reward_name.text().lower()

        for row_data in rewards:
            recomendador, nuevo_alumno, nombre_recompensa = row_data[1], row_data[2], row_data[5]
            if (filter_recommender and filter_recommender not in recomendador.lower()) or \
               (filter_new_student and filter_new_student not in nuevo_alumno.lower()) or \
               (filter_reward_name and filter_reward_name not in nombre_recompensa.lower()):
                continue

            row = self.table.rowCount()
            self.table.insertRow(row)
            self.table.setItem(row, 0, QTableWidgetItem(recomendador))
            self.table.setItem(row, 1, QTableWidgetItem(nuevo_alumno))
            self.table.setItem(row, 2, QTableWidgetItem(nombre_recompensa))
            self.table.setItem(row, 3, QTableWidgetItem(str(row_data[3])))
            self.table.setItem(row, 4, QTableWidgetItem(row_data[4]))

    def add_reward(self):
        """Añade una nueva recompensa con validaciones"""
        recommender_id = self.recommender_combo.currentData()
        new_student_id = self.new_student_combo.currentData()
        months_text = self.months_input.text().strip()

        # Validar que los IDs no estén vacíos
        if recommender_id is None or new_student_id is None:
            QMessageBox.warning(self, "Error", "Selecciona un alumno recomendador y uno nuevo.")
            return

        # Validar que sean distintos
        if recommender_id == new_student_id:
            QMessageBox.warning(self, "Error", "Un alumno no puede recomendarse a sí mismo.")
            return

        # Validar que se ingresó una cantidad de meses válida (1, 3 o 6)
        if not months_text.isdigit() or int(months_text) not in [1, 3, 6]:
            QMessageBox.warning(self, "Error", "Solo se permiten recompensas de 1, 3 o 6 meses.")
            return

        months = int(months_text)

        # Validar que no se haya otorgado ya una recompensa al nuevo alumno
        if reward_already_granted(new_student_id):
            QMessageBox.warning(self, "Error", "Este alumno ya tiene una recompensa asignada.")
            return

        # Insertar recompensa
        date_awarded = self.date_input.text().strip()

        # Validar formato de fecha
        try:
            datetime.strptime(date_awarded, "%Y-%m-%d")
        except ValueError:
            QMessageBox.warning(self, "Error", "La fecha debe estar en formato YYYY-MM-DD.")
            return

        insert_reward(recommender_id, new_student_id, months, date_awarded)

        QMessageBox.information(self, "Éxito", "Recompensa añadida correctamente.")

        # Recargar tabla y limpiar campos
        self.load_rewards()
        self.months_input.clear()
        self.date_input.clear()

    def delete_reward(self):
        """Elimina la recompensa seleccionada"""
        selected = self.table.currentRow()
        if selected == -1:
            QMessageBox.warning(self, "Error", "Selecciona una recompensa para eliminar.")
            return

        recommender = self.table.item(selected, 0).text()
        new_student = self.table.item(selected, 1).text()

        confirm = QMessageBox.question(
            self,
            "Confirmar Borrado",
            f"¿Estás seguro de borrar la recompensa de '{recommender}' a '{new_student}'?",
            QMessageBox.Yes | QMessageBox.No
        )

        if confirm == QMessageBox.Yes:
            try:
                from db.database import delete_reward  # Asegúrate de que existe esta función
                delete_reward(recommender, new_student)
                QMessageBox.information(self, "Éxito", "Recompensa eliminada correctamente.")
                self.load_rewards()
            except Exception as e:
                QMessageBox.critical(self, "Error", f"No se pudo eliminar la recompensa:\n{e}")

    def clear_filters(self):
        """Limpia los campos de búsqueda"""
        self.search_recommender.clear()
        self.search_new_student.clear()
        self.search_reward_name.clear()
        self.load_rewards()

    def show_expiring_alerts(self):
        """Muestra una alerta con los alumnos cuyo bono está por terminar"""
        expiring = get_students_with_expiring_bonus()
        if not expiring:
            QMessageBox.information(self, "Sin alertas", "No hay bonos que finalicen en los próximos 7 días.")
            return

        message = "Bonos que terminan pronto:\n\n"
        for name, end_date, days_left in expiring:
            message += f"{name} → termina el {end_date} (en {days_left} días)\n"

        QMessageBox.information(self, "Bonos por finalizar", message)

    def export_to_csv(self):
        """Exporta las reconpensas a un archivo CSV"""
        # Abrir diálogo para elegir dónde guardar el archivo
        path, _ = QFileDialog.getSaveFileName(self, "Guardar como CSV", "recompensas.csv", "CSV Files (*.csv)")
        if not path:
            return

        # Por si se cambia el nombre y no se le pone la extension .csv
        if not path.endswith(".csv"):
            path += ".csv"

        with open(path, mode='w', newline='', encoding='utf-8') as file:
            writer = csv.writer(file)
            writer.writerow(["Recomendador", "Alumno nuevo", "Nombre de la recompensa", "Meses de recompensa", "Fecha de inicio bono"])

            for row in range(self.table.rowCount()):
                row_data = [self.table.item(row, col).text() if self.table.item(row, col) else "" for col in
                            range(self.table.columnCount())]
                writer.writerow(row_data)

        QMessageBox.information(self, "Éxito", f"Archivo CSV exportado correctamente:\n{path}")

    def export_to_excel(self):
         """Exporta las recompensas a un archivo Excel (.xlsx) usando openpyxl"""
         # Abrir diálogo para elegir dónde guardar el archivo
         path, _ = QFileDialog.getSaveFileName(self, "Guardar como Excel", "recompensas.xlsx", "Archivos Excel (*.xlsx)")
         if not path:
             return  # Cancelado por el usuario

         # Por si se cambia el nombre y no se le pone la extension .xlsx
         if not path.endswith(".xlsx"):
             path += ".xlsx"

         workbook = openpyxl.Workbook()
         sheet = workbook.active
         sheet.title = "Recompensas"

         # Escribir cabeceras
         headers = ["Recomendador", "Alumno nuevo", "Nombre de la recompensa", "Meses de recompensa", "Fecha de inicio bono"]
         for col, header in enumerate(headers, start=1):
             sheet.cell(row=1, column=col, value=header)

        # Escribir datos de la tabla
         for row in range(self.table.rowCount()):
             for col in range(self.table.columnCount()):
                 item = self.table.item(row, col)
                 sheet.cell(row=row + 2, column=col + 1, value=item.text() if item else "") # +2 porque empieza en fila 2

         workbook.save(path)
         QMessageBox.information(self, "Éxito", f"Archivo Excel exportado correctamente:\n{path}")